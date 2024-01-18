import os
import openpyxl
import requests
import base64
from bs4 import BeautifulSoup
from io import BytesIO
from PIL import Image


folder_name = 'Images'
os.makedirs(folder_name, exist_ok=True)

workbook = openpyxl.load_workbook('./test.xlsx')
sheet = workbook.active

rows = []
header = [cell.value for cell in next(sheet.iter_rows())]
for row in sheet.iter_rows(min_row=2):
    row_data = {header[i]: cell.value for i, cell in enumerate(row)}
    rows.append(row_data)

for row in rows:
    name = row['Name']
    rollno = row['Rno']
    url = row['Url']
    filename = f"{name}_{rollno}.png"  
    file_path = os.path.join(folder_name, filename)

    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        img_element = soup.find('img')

        if img_element:
            base64_data = img_element['src'].split(',')[1]
            decoded_data = base64.b64decode(base64_data)
            image_stream = BytesIO(decoded_data)
            image = Image.open(image_stream)

            os.makedirs(folder_name, exist_ok=True)

            file_path = os.path.join(folder_name, filename)

            image.save(file_path)

            print(f"Downloaded {filename}")
        else:
            print("No image found on the webpage.")
    else:
        print(f"Failed to download image from {url}")
